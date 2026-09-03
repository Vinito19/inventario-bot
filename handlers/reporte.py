from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, CallbackQueryHandler, CommandHandler, ConversationHandler, MessageHandler, filters

from database import (
    obtener_resumen,
    esta_registrado,
    es_admin,
    obtener_resumen_ventas,
    obtener_ventas,
    obtener_cambios,
    obtener_ventas_por_fecha,
    obtener_resumen_ventas_por_fecha,
    obtener_cambios_por_fecha,
)
from handlers.utils import edit_mensaje, guardar_mensaje, borrar_mensajes, finalizar
from keyboards import botones_volver
from excel_export import generar_excel, generar_excel_ventas, generar_excel_cambios

VENTAS_MES, CAMBIOS_MES = range(2)


async def start_reporte(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not esta_registrado(user_id):
        await update.message.reply_text("No tienes acceso al bot.")
        return

    resumen = obtener_resumen()
    admin = es_admin(user_id)

    texto = (
        "📊 REPORTE DE INVENTARIO\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 Total artículos:    {resumen['total']}\n"
        f"🔢 Total unidades:     {resumen['unidades']}\n"
        f"🚫 Stock en cero:      {resumen['cero']}\n"
        f"💰 Valor total:        ${resumen['valor']:,.2f}\n\n"
    )

    if resumen["por_categoria"]:
        texto += "📂 Por categoría:\n"
        for cat in resumen["por_categoria"]:
            texto += f"  • {cat['nombre']}: {cat['total']} artículos · {cat['unidades']} unidades\n"

    if resumen["sin_categoria"]:
        texto += "\n📌 Sin categoría:\n"
        for s in resumen["sin_categoria"]:
            texto += f"  • {s['total']} artículos · {s['unidades']} unidades\n"

    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

    keyboard = []

    if admin:
        keyboard.append([InlineKeyboardButton("📊 Exportar a Excel", callback_data="exportar_excel")])
        keyboard.append([InlineKeyboardButton("💵 Ver ventas", callback_data="ver_ventas")])
        keyboard.append([InlineKeyboardButton("📅 Ventas por mes", callback_data="ver_ventas_mes")])
        keyboard.append([InlineKeyboardButton("🛠️ Ver cambios", callback_data="ver_cambios")])
        keyboard.append([InlineKeyboardButton("📅 Cambios por mes", callback_data="ver_cambios_mes")])

    keyboard.append([InlineKeyboardButton("🧹 Ver artículos en cero", callback_data="ver_stock_cero")])

    if admin:
        keyboard.append([InlineKeyboardButton("🧹 Limpiar artículos en cero", callback_data="limpiar")])

    keyboard.append([InlineKeyboardButton("🏠 Volver al menú", callback_data="inicio")])

    msg = await update.message.reply_text(texto, reply_markup=InlineKeyboardMarkup(keyboard))
    guardar_mensaje(update, context, msg)


async def callback_reporte(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not esta_registrado(user_id):
        await edit_mensaje(query, "❌ No tienes acceso al bot.")
        return

    resumen = obtener_resumen()
    admin = es_admin(user_id)

    texto = (
        "📊 REPORTE DE INVENTARIO\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 Total artículos:    {resumen['total']}\n"
        f"🔢 Total unidades:     {resumen['unidades']}\n"
        f"🚫 Stock en cero:      {resumen['cero']}\n"
        f"💰 Valor total:        ${resumen['valor']:,.2f}\n\n"
    )

    if resumen["por_categoria"]:
        texto += "📂 Por categoría:\n"
        for cat in resumen["por_categoria"]:
            texto += f"  • {cat['nombre']}: {cat['total']} artículos · {cat['unidades']} unidades\n"

    if resumen["sin_categoria"]:
        texto += "\n📌 Sin categoría:\n"
        for s in resumen["sin_categoria"]:
            texto += f"  • {s['total']} artículos · {s['unidades']} unidades\n"

    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

    keyboard = []

    if admin:
        keyboard.append([InlineKeyboardButton("📊 Exportar a Excel", callback_data="exportar_excel")])
        keyboard.append([InlineKeyboardButton("💵 Ver ventas", callback_data="ver_ventas")])
        keyboard.append([InlineKeyboardButton("📅 Ventas por mes", callback_data="ver_ventas_mes")])
        keyboard.append([InlineKeyboardButton("🛠️ Ver cambios", callback_data="ver_cambios")])
        keyboard.append([InlineKeyboardButton("📅 Cambios por mes", callback_data="ver_cambios_mes")])

    keyboard.append([InlineKeyboardButton("🧹 Ver artículos en cero", callback_data="ver_stock_cero")])

    if admin:
        keyboard.append([InlineKeyboardButton("🧹 Limpiar artículos en cero", callback_data="limpiar")])

    keyboard.append([InlineKeyboardButton("🏠 Volver al menú", callback_data="inicio")])

    await edit_mensaje(query, texto, reply_markup=InlineKeyboardMarkup(keyboard))


async def exportar_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "Solo el administrador puede exportar a Excel.")
        return

    await borrar_mensajes(context, chat_id=user_id)
    await edit_mensaje(query, "Generando archivo Excel...")

    try:
        archivo = generar_excel()
        with open(archivo, "rb") as f:
            doc_msg = await context.bot.send_document(
                chat_id=user_id,
                document=f,
                caption="Inventario exportado correctamente",
            )
        guardar_mensaje(update, context, doc_msg)
        await edit_mensaje(query, "Archivo Excel enviado!", reply_markup=botones_volver())
    except Exception as e:
        await edit_mensaje(
            query,
            f"Error al generar Excel: {str(e)}",
            reply_markup=botones_volver(),
        )
    finally:
        try:
            import os
            if os.path.exists(archivo):
                os.remove(archivo)
        except Exception:
            pass


async def ver_stock_cero(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    from database import obtener_stock_cero
    stock_cero = obtener_stock_cero()

    if not stock_cero:
        await edit_mensaje(
            query,
            "🧹 No hay artículos con stock en cero.",
            reply_markup=botones_volver(),
        )
        return

    texto = "🚫 ARTÍCULOS SIN STOCK\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    for r in stock_cero[:15]:
        texto += f"• {r['codigo']} - {r['nombre']}\n"

    if len(stock_cero) > 15:
        texto += f"\n... y {len(stock_cero) - 15} más\n"

    texto += f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\nTotal: {len(stock_cero)} artículos"

    await edit_mensaje(query, texto, reply_markup=botones_volver())


async def ver_ventas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "❌ Solo el administrador puede ver las ventas.")
        return

    resumen = obtener_resumen_ventas()
    ventas = obtener_ventas()

    texto = (
        "💵 REPORTE DE VENTAS\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🧾 Número de ventas:   {resumen['ventas']}\n"
        f"📦 Unidades vendidas:  {resumen['unidades']}\n"
        f"💰 Total vendido:      ${resumen['total']:,.2f}\n\n"
    )

    if ventas:
        texto += "📋 Últimas ventas:\n"
        for v in ventas[:10]:
            texto += (
                f"• {v['fecha'][:16]} | {v['codigo']} - {v['nombre']}\n"
                f"  {v['cantidad']} und × ${v['precio_unitario']:.2f} × subtotal ${v['subtotal']:.2f} × {v['usuario_nombre']}\n"
            )
        if len(ventas) > 10:
            texto += f"\n... y {len(ventas) - 10} más\n"

    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

    keyboard = [
        [InlineKeyboardButton("📊 Exportar ventas a Excel", callback_data="exportar_ventas")],
        [InlineKeyboardButton("🏠 Volver al menú", callback_data="inicio")],
    ]
    await edit_mensaje(query, texto, reply_markup=InlineKeyboardMarkup(keyboard))


async def exportar_ventas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "Solo el administrador puede exportar ventas.")
        return

    await borrar_mensajes(context, chat_id=user_id)
    await edit_mensaje(query, "Generando archivo de ventas...")

    try:
        archivo = generar_excel_ventas()
        with open(archivo, "rb") as f:
            doc_msg = await context.bot.send_document(
                chat_id=user_id,
                document=f,
                caption="Reporte de ventas",
            )
        guardar_mensaje(update, context, doc_msg)
        await edit_mensaje(query, "Archivo de ventas enviado!", reply_markup=botones_volver())
    except Exception as e:
        await edit_mensaje(query, f"Error al generar ventas: {str(e)}", reply_markup=botones_volver())
    finally:
        try:
            import os
            if os.path.exists(archivo):
                os.remove(archivo)
        except Exception:
            pass


async def ver_cambios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "❌ Solo el administrador puede ver los cambios.")
        return

    cambios = obtener_cambios()

    texto = (
        "🛠️ REPORTE DE CAMBIOS\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    )

    if not cambios:
        texto += "No hay cambios registrados."
    else:
        texto += f"Total de cambios: {len(cambios)}\n\n"
        texto += "📋 Últimos cambios:\n"
        for c in cambios[:10]:
            texto += (
                f"• {c['fecha'][:16]} | {c['repuesto_codigo']}\n"
                f"  {c['campo']}: '{c['valor_anterior']}' → '{c['valor_nuevo']}'\n"
                f"  👤 {c['usuario_nombre']}\n"
            )
        if len(cambios) > 10:
            texto += f"\n... y {len(cambios) - 10} más\n"

    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

    keyboard = [
        [InlineKeyboardButton("📊 Exportar cambios a Excel", callback_data="exportar_cambios")],
        [InlineKeyboardButton("🏠 Volver al menú", callback_data="inicio")],
    ]
    await edit_mensaje(query, texto, reply_markup=InlineKeyboardMarkup(keyboard))


async def exportar_cambios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "❌ Solo el administrador puede exportar cambios.")
        return

    await borrar_mensajes(context, chat_id=user_id)
    await edit_mensaje(query, "⏳ Generando archivo de cambios...")

    try:
        archivo = generar_excel_cambios()
        with open(archivo, "rb") as f:
            doc_msg = await context.bot.send_document(
                chat_id=user_id,
                document=f,
                caption="🛠️ Reporte de cambios de edición",
            )
        guardar_mensaje(update, context, doc_msg)
        await edit_mensaje(query, "✅ Archivo de cambios enviado!", reply_markup=botones_volver())
    except Exception as e:
        await edit_mensaje(query, f"❌ Error al generar cambios: {str(e)}", reply_markup=botones_volver())
    finally:
        try:
            import os
            if os.path.exists(archivo):
                os.remove(archivo)
        except Exception:
            pass


# ---------- FILTRO POR MES ----------

async def ver_ventas_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "❌ Solo el administrador puede ver ventas por mes.")
        return

    await edit_mensaje(
        query,
        "📅 VENTAS POR MES\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Escribe el mes y año (formato: MM/AAAA o AAAA-MM):\n"
        "Ejemplos: 09/2026  o  2026-09",
        reply_markup=botones_volver(),
    )
    return VENTAS_MES


async def ventas_mes_recibido(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()

    # Parsear mes/anio
    try:
        if "/" in texto:
            mes, anio = texto.split("/")
        elif "-" in texto:
            anio, mes = texto.split("-")
        else:
            raise ValueError

        mes = int(mes)
        anio = int(anio)
        if not (1 <= mes <= 12):
            raise ValueError
        if not (2020 <= anio <= 2030):
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "⚠️ Formato inválido. Usa MM/AAAA o AAAA-MM (ej: 09/2026 o 2026-09):",
            reply_markup=botones_volver(),
        )
        return VENTAS_MES

    # Calcular rango de fechas
    from datetime import datetime, timedelta
    fecha_inicio = f"{anio:04d}-{mes:02d}-01"
    if mes == 12:
        fecha_fin = f"{anio+1:04d}-01-01"
    else:
        fecha_fin = f"{anio:04d}-{mes+1:02d}-01"

    resumen = obtener_resumen_ventas_por_fecha(fecha_inicio, fecha_fin)
    ventas = obtener_ventas_por_fecha(fecha_inicio, fecha_fin)

    mes_nombre = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ][mes]

    texto = (
        f"📅 VENTAS DE {mes_nombre.upper()} {anio}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🧾 Número de ventas:   {resumen['ventas']}\n"
        f"📦 Unidades vendidas:  {resumen['unidades']}\n"
        f"💰 Total vendido:      ${resumen['total']:,.2f}\n\n"
    )

    if ventas:
        texto += "📋 Ventas:\n"
        for v in ventas:
            texto += (
                f"• {v['fecha'][:16]} | {v['codigo']} - {v['nombre']}\n"
                f"  {v['cantidad']} und × ${v['precio_unitario']:.2f} × subtotal ${v['subtotal']:.2f} × {v['usuario_nombre']}\n"
            )

    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

    keyboard = [
        [InlineKeyboardButton("📊 Exportar a Excel", callback_data=f"exportar_ventas_mes_{anio}_{mes:02d}")],
        [InlineKeyboardButton("📅 Otro mes", callback_data="ver_ventas_mes")],
        [InlineKeyboardButton("🏠 Volver al menú", callback_data="inicio")],
    ]
    await update.message.reply_text(texto, reply_markup=InlineKeyboardMarkup(keyboard))

    return ConversationHandler.END


async def ver_cambios_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "❌ Solo el administrador puede ver cambios por mes.")
        return

    await edit_mensaje(
        query,
        "📅 CAMBIOS POR MES\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Escribe el mes y año (formato: MM/AAAA o AAAA-MM):\n"
        "Ejemplos: 09/2026  o  2026-09",
        reply_markup=botones_volver(),
    )
    return CAMBIOS_MES


async def cambios_mes_recibido(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()

    try:
        if "/" in texto:
            mes, anio = texto.split("/")
        elif "-" in texto:
            anio, mes = texto.split("-")
        else:
            raise ValueError

        mes = int(mes)
        anio = int(anio)
        if not (1 <= mes <= 12):
            raise ValueError
        if not (2020 <= anio <= 2030):
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "⚠️ Formato inválido. Usa MM/AAAA o AAAA-MM (ej: 09/2026 o 2026-09):",
            reply_markup=botones_volver(),
        )
        return CAMBIOS_MES

    from datetime import datetime
    if mes == 12:
        fecha_fin = f"{anio+1:04d}-01-01"
    else:
        fecha_fin = f"{anio:04d}-{mes+1:02d}-01"
    fecha_inicio = f"{anio:04d}-{mes:02d}-01"

    cambios = obtener_cambios_por_fecha(fecha_inicio, fecha_fin)

    mes_nombre = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ][mes]

    texto = (
        f"📅 CAMBIOS DE {mes_nombre.upper()} {anio}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🛠️ Total de cambios: {len(cambios)}\n\n"
    )

    if cambios:
        texto += "📋 Cambios:\n"
        for c in cambios:
            texto += (
                f"• {c['fecha'][:16]} | {c['repuesto_codigo']}\n"
                f"  {c['campo']}: '{c['valor_anterior']}' → '{c['valor_nuevo']}'\n"
                f"  👤 {c['usuario_nombre']}\n"
            )

    texto += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"

    keyboard = [
        [InlineKeyboardButton("📊 Exportar a Excel", callback_data=f"exportar_cambios_mes_{anio}_{mes:02d}")],
        [InlineKeyboardButton("📅 Otro mes", callback_data="ver_cambios_mes")],
        [InlineKeyboardButton("🏠 Volver al menú", callback_data="inicio")],
    ]
    await update.message.reply_text(texto, reply_markup=InlineKeyboardMarkup(keyboard))

    return ConversationHandler.END


async def exportar_ventas_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "Solo el administrador puede exportar.")
        return

    # callback_data: exportar_ventas_mes_YYYY_MM
    try:
        _, _, anio, mes = query.data.split("_")
        anio = int(anio)
        mes = int(mes)
    except Exception:
        await edit_mensaje(query, "Error en los parametros.")
        return

    if mes == 12:
        fecha_fin = f"{anio+1:04d}-01-01"
    else:
        fecha_fin = f"{anio:04d}-{mes+1:02d}-01"
    fecha_inicio = f"{anio:04d}-{mes:02d}-01"

    await borrar_mensajes(context, chat_id=user_id)
    await edit_mensaje(query, "Generando archivo...")

    try:
        from database import obtener_ventas_por_fecha
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ventas = obtener_ventas_por_fecha(fecha_inicio, fecha_fin)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Ventas_{anio}_{mes:02d}"

        headers = ["Fecha", "Codigo", "Nombre", "Cantidad", "Precio reg.", "Precio unit.", "Desc/Rec", "Total", "Usuario"]
        from excel_export import _estilos_excel
        _estilos_excel(wb, ws, headers)

        from openpyxl.styles import Alignment
        for row_num, v in enumerate(ventas, 2):
            ws.cell(row=row_num, column=1, value=v["fecha"])
            ws.cell(row=row_num, column=2, value=v["codigo"])
            ws.cell(row=row_num, column=3, value=v["nombre"])
            ws.cell(row=row_num, column=4, value=v["cantidad"])
            ws.cell(row=row_num, column=5, value=v["precio_registrado"])
            ws.cell(row=row_num, column=6, value=v["precio_unitario"])
            ws.cell(row=row_num, column=7, value=round(v["precio_unitario"] - v["precio_registrado"], 2))
            ws.cell(row=row_num, column=8, value=v["subtotal"])
            ws.cell(row=row_num, column=9, value=v["usuario_nombre"])
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
            for col in (5, 6, 7, 8):
                ws.cell(row=row_num, column=col).number_format = '#,##0.00'
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="right")

        widths = {"A": 20, "B": 15, "C": 30, "D": 12, "E": 17, "F": 17, "G": 18, "H": 15, "I": 20}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        import os
        from datetime import datetime
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo = f"ventas_{anio}_{mes:02d}_{fecha_str}.xlsx"
        ruta = os.path.join(os.getcwd(), archivo)
        wb.save(ruta)

        with open(ruta, "rb") as f:
            doc_msg = await context.bot.send_document(
                chat_id=user_id,
                document=f,
                caption=f"Ventas {anio}-{mes:02d}",
            )
        guardar_mensaje(update, context, doc_msg)
        await edit_mensaje(query, "Archivo enviado!", reply_markup=botones_volver())
        os.remove(ruta)
    except Exception as e:
        await edit_mensaje(query, f"Error: {str(e)}", reply_markup=botones_volver())


async def exportar_cambios_mes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not es_admin(user_id):
        await edit_mensaje(query, "Solo el administrador puede exportar.")
        return

    try:
        _, _, anio, mes = query.data.split("_")
        anio = int(anio)
        mes = int(mes)
    except Exception:
        await edit_mensaje(query, "Error en los parametros.")
        return

    if mes == 12:
        fecha_fin = f"{anio+1:04d}-01-01"
    else:
        fecha_fin = f"{anio:04d}-{mes+1:02d}-01"
    fecha_inicio = f"{anio:04d}-{mes:02d}-01"

    await borrar_mensajes(context, chat_id=user_id)
    await edit_mensaje(query, "Generando archivo...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        cambios = obtener_cambios_por_fecha(fecha_inicio, fecha_fin)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Cambios_{anio}_{mes:02d}"

        headers = ["Fecha", "Codigo repuesto", "Campo", "Valor anterior", "Valor nuevo", "Usuario"]
        from excel_export import _estilos_excel
        _estilos_excel(wb, ws, headers)

        for row_num, c in enumerate(cambios, 2):
            ws.cell(row=row_num, column=1, value=c["fecha"])
            ws.cell(row=row_num, column=2, value=c["repuesto_codigo"])
            ws.cell(row=row_num, column=3, value=c["campo"])
            ws.cell(row=row_num, column=4, value=c["valor_anterior"])
            ws.cell(row=row_num, column=5, value=c["valor_nuevo"])
            ws.cell(row=row_num, column=6, value=c["usuario_nombre"])

        widths = {"A": 20, "B": 18, "C": 18, "D": 25, "E": 25, "F": 20}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        import os
        from datetime import datetime
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo = f"cambios_{anio}_{mes:02d}_{fecha_str}.xlsx"
        ruta = os.path.join(os.getcwd(), archivo)
        wb.save(ruta)

        with open(ruta, "rb") as f:
            doc_msg = await context.bot.send_document(
                chat_id=user_id,
                document=f,
                caption=f"Cambios {anio}-{mes:02d}",
            )
        guardar_mensaje(update, context, doc_msg)
        await edit_mensaje(query, "Archivo enviado!", reply_markup=botones_volver())
        os.remove(ruta)
    except Exception as e:
        await edit_mensaje(query, f"Error: {str(e)}", reply_markup=botones_volver())


reporte_handler = CommandHandler("reporte", start_reporte)
reporte_callback_handler = CallbackQueryHandler(callback_reporte, pattern="^reporte$")
exportar_excel_handler = CallbackQueryHandler(exportar_excel, pattern="^exportar_excel$")
ver_stock_cero_handler = CallbackQueryHandler(ver_stock_cero, pattern="^ver_stock_cero$")
ver_ventas_handler = CallbackQueryHandler(ver_ventas, pattern="^ver_ventas$")
exportar_ventas_handler = CallbackQueryHandler(exportar_ventas, pattern="^exportar_ventas$")
ver_cambios_handler = CallbackQueryHandler(ver_cambios, pattern="^ver_cambios$")
exportar_cambios_handler = CallbackQueryHandler(exportar_cambios, pattern="^exportar_cambios$")

ver_ventas_mes_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(ver_ventas_mes, pattern="^ver_ventas_mes$")],
    states={VENTAS_MES: [MessageHandler(filters.TEXT & ~filters.COMMAND, ventas_mes_recibido)]},
    fallbacks=[CallbackQueryHandler(finalizar, pattern="^(inicio|cancelar)$")],
)

ver_cambios_mes_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(ver_cambios_mes, pattern="^ver_cambios_mes$")],
    states={CAMBIOS_MES: [MessageHandler(filters.TEXT & ~filters.COMMAND, cambios_mes_recibido)]},
    fallbacks=[CallbackQueryHandler(finalizar, pattern="^(inicio|cancelar)$")],
)

exportar_ventas_mes_handler = CallbackQueryHandler(exportar_ventas_mes, pattern=r"^exportar_ventas_mes_\d{4}_\d{2}$")
exportar_cambios_mes_handler = CallbackQueryHandler(exportar_cambios_mes, pattern=r"^exportar_cambios_mes_\d{4}_\d{2}$")