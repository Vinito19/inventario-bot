import os
import sys
import zipfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import pytest
import database
from database import agregar_categoria, agregar_repuesto


@pytest.fixture
def repuestos_en_db():
    cat_id = agregar_categoria("Frenos")
    fotos = [f"f{i}" for i in range(1, 5)]
    agregar_repuesto("XL-001", "Pastilla", "d1", 0, 10.0, fotos, cat_id, "A1")
    agregar_repuesto("XL-002", "Disco", "d2", 3, 20.0, fotos, cat_id, "A2")
    agregar_repuesto("XL-003", "Filtro", "d3", 8, 5.0, fotos, cat_id, "B1")
    return cat_id


def test_generar_excel(repuestos_en_db):
    from excel_export import generar_excel

    ruta = generar_excel()
    try:
        assert os.path.exists(ruta)
        assert ruta.endswith(".xlsx")
        assert os.path.getsize(ruta) > 0
    finally:
        if os.path.exists(ruta):
            os.remove(ruta)


def test_excel_contenido(repuestos_en_db):
    """Valida que el Excel contenga los datos esperados."""
    from openpyxl import load_workbook
    from excel_export import generar_excel

    ruta = generar_excel()
    try:
        wb = load_workbook(ruta)
        ws = wb.active
        assert ws.title == "Inventario"
        # Cabecera
        assert ws.cell(row=1, column=1).value == "Código"
        # Filas de datos (3 repuestos + cabecera)
        assert ws.max_row >= 4
        codigos = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "XL-001" in codigos
        assert "XL-002" in codigos
        assert "XL-003" in codigos
    finally:
        if os.path.exists(ruta):
            os.remove(ruta)


def test_hacer_backup_db():
    from backup import hacer_backup_db

    ruta = hacer_backup_db()
    try:
        assert ruta and os.path.exists(ruta)
        assert os.path.getsize(ruta) > 0
    finally:
        if ruta and os.path.exists(ruta):
            os.remove(ruta)


def test_backup_db_valido():
    """El backup debe ser una base SQLite abrible."""
    import sqlite3
    from backup import hacer_backup_db

    ruta = hacer_backup_db()
    try:
        conn = sqlite3.connect(ruta)
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tablas = [row[0] for row in cur.fetchall()]
        assert "repuestos" in tablas
        assert "usuarios" in tablas
        conn.close()
    finally:
        if ruta and os.path.exists(ruta):
            os.remove(ruta)


def test_nombre_backup_extension():
    from backup import nombre_backup

    nombre = nombre_backup()
    assert nombre.endswith(".db")
    assert "backup_inventario" in nombre