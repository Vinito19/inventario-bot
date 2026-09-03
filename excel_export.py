import os
from datetime import datetime

from openpyxl import Workbook

from database import get_connection


def generar_excel():
    conn = get_connection()
    try:
        cursor = conn.cursor()

        cursor.execute("""
            SELECT r.codigo, r.nombre, r.descripcion, r.cantidad, r.precio,
                   r.ubicacion, r.fecha, c.nombre as categoria
            FROM repuestos r
            LEFT JOIN categorias c ON r.categoria_id = c.id
            ORDER BY r.codigo
        """)

        repuestos = cursor.fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["Código", "Nombre", "Descripción", "Categoría", "Cantidad", "Precio", "Ubicación", "Fecha"]

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, r in enumerate(repuestos, 2):
        ws.cell(row=row_num, column=1, value=r["codigo"]).border = thin_border
        ws.cell(row=row_num, column=2, value=r["nombre"]).border = thin_border
        ws.cell(row=row_num, column=3, value=r["descripcion"]).border = thin_border
        ws.cell(row=row_num, column=4, value=r["categoria"] or "Sin categoría").border = thin_border
        ws.cell(row=row_num, column=5, value=r["cantidad"]).border = thin_border
        ws.cell(row=row_num, column=6, value=r["precio"]).border = thin_border
        ws.cell(row=row_num, column=7, value=r["ubicacion"]).border = thin_border
        ws.cell(row=row_num, column=8, value=r["fecha"]).border = thin_border

        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="right")

        if r["cantidad"] == 0:
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = red_fill
        elif r["cantidad"] < 5:
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = yellow_fill

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo = f"inventario_{fecha_str}.xlsx"
    ruta = os.path.join(os.getcwd(), archivo)
    wb.save(ruta)

    return ruta


def _estilos_excel(wb, ws, headers):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def generar_excel_ventas():
    from database import obtener_ventas

    ventas = obtener_ventas()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ["Fecha", "Código", "Nombre", "Cantidad", "Precio unitario", "Precio registrado", "Subtotal", "Usuario"]
    _estilos_excel(wb, ws, headers)

    from openpyxl.styles import Alignment

    for row_num, v in enumerate(ventas, 2):
        ws.cell(row=row_num, column=1, value=v["fecha"])
        ws.cell(row=row_num, column=2, value=v["codigo"])
        ws.cell(row=row_num, column=3, value=v["nombre"])
        ws.cell(row=row_num, column=4, value=v["cantidad"])
        ws.cell(row=row_num, column=5, value=v["precio_unitario"])
        ws.cell(row=row_num, column=6, value=v["precio_registrado"])
        ws.cell(row=row_num, column=7, value=v["subtotal"])
        ws.cell(row=row_num, column=8, value=v["usuario_nombre"])

        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
        for col in (5, 6, 7):
            ws.cell(row=row_num, column=col).number_format = '#,##0.00'
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="right")

    widths = {"A": 20, "B": 15, "C": 30, "D": 12, "E": 15, "F": 17, "G": 15, "H": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo = f"ventas_{fecha_str}.xlsx"
    ruta = os.path.join(os.getcwd(), archivo)
    wb.save(ruta)
    return ruta


def generar_excel_cambios():
    from database import obtener_cambios

    cambios = obtener_cambios()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cambios"

    headers = ["Fecha", "Código repuesto", "Campo", "Valor anterior", "Valor nuevo", "Usuario"]
    _estilos_excel(wb, ws, headers)

    for row_num, c in enumerate(cambios, 2):
        ws.cell(row=row_num, column=1, value=c["fecha"])
        ws.cell(row=row_num, column=2, value=c["repuesto_codigo"])
        ws.cell(row=row_num, column=3, value=c["campo"])
        ws.cell(row=row_num, column=4, value=c["valor_anterior"])
        ws.cell(row=row_num, column=5, value=c["valor_nuevo"])
        ws.cell(row=row_num, column=6, value=c["usuario_nombre"])

    widths = {"A": 20, "B": 18, "C": 18, "D": 25, "E": 25, "F": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo = f"cambios_{fecha_str}.xlsx"
    ruta = os.path.join(os.getcwd(), archivo)
    wb.save(ruta)
    return ruta